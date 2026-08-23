import base64
import io
import logging
import os
import re
from datetime import datetime

import pandas as pd
import plotly.express as px
import streamlit as st

from news_monitor import NewsClassifier, NewsPipeline, NewsRepository
from news_monitor.config import LOG_PATH, LOGO_PATH, START_YEAR, UI_COLUMNS

MONTH_NAMES_ID = (
    "",
    "Januari",
    "Februari",
    "Maret",
    "April",
    "Mei",
    "Juni",
    "Juli",
    "Agustus",
    "September",
    "Oktober",
    "November",
    "Desember",
)

logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)

st.set_page_config(
    page_title="Monitoring Berita Ekonomi Lamongan - BPS",
    page_icon=str(LOGO_PATH) if LOGO_PATH.exists() else "📰",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
main { background-color: #f8fafc; }
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
[data-testid="stElementToolbar"] button[title="Download as CSV"],
[data-testid="stElementToolbar"] button[aria-label="Download as CSV"],
[data-testid="stElementToolbar"] button:has(svg path[d*="M19 9h-4V3H9v6H5l7 7 7-7"]) {
    display: none !important;
}
.dashboard-header {
    background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%);
    padding: 22px 28px; border-radius: 16px; color: white; margin-bottom: 25px;
    box-shadow: 0 4px 12px rgba(30, 58, 138, 0.15); display: flex;
    align-items: center; gap: 20px;
}
.dashboard-logo {
    width: 75px; height: 75px; object-fit: contain; background: white;
    padding: 8px; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.2);
    flex-shrink: 0;
}
.dashboard-title { font-size: 26px; font-weight: 800; margin: 0; color: white; }
.dashboard-subtitle { font-size: 14px; color: #e0f2fe; margin-top: 7px; }
.section-header {
    font-size: 18px; font-weight: 700; color: #1e293b; margin-top: 15px;
    margin-bottom: 15px; border-left: 4px solid #2563eb; padding-left: 10px;
}
</style>
""",
    unsafe_allow_html=True,
)


@st.cache_resource
def get_repository() -> NewsRepository:
    return NewsRepository()


def get_api_key() -> str | None:
    try:
        value = (
            st.secrets.get("GEMINI_API_KEY")
            or st.secrets.get("GOOGLE_API_KEY")
            or os.getenv("GEMINI_API_KEY")
            or os.getenv("GOOGLE_API_KEY")
        )
    except Exception:
        value = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    return str(value).strip() if value else None


def has_valid_api_key_format() -> bool:
    api_key = get_api_key()
    return bool(
        api_key
        and api_key.startswith(("AIza", "AQ."))
        and len(api_key) >= 35
    )


def create_empty_data() -> pd.DataFrame:
    return pd.DataFrame(columns=UI_COLUMNS)


def load_data(year: int) -> pd.DataFrame:
    frame = get_repository().load_dataframe(year)
    for column in UI_COLUMNS:
        if column not in frame.columns:
            frame[column] = ""
    return frame


def build_classifier() -> NewsClassifier:
    api_key = get_api_key() if has_valid_api_key_format() else None
    return NewsClassifier(api_key=api_key)


def is_quota_error(error: str) -> bool:
    normalized = str(error).casefold()
    return any(
        marker in normalized
        for marker in ("429", "quota", "rate limit", "too_many_requests")
    )


def quota_wait_seconds(error: str) -> int | None:
    match = re.search(r"retry in ([0-9.]+)s", str(error), re.IGNORECASE)
    return max(1, round(float(match.group(1)))) if match else None


def show_quota_warning(error: str) -> None:
    wait_seconds = quota_wait_seconds(error)
    wait_text = (
        f" Coba lagi sekitar {wait_seconds} detik."
        if wait_seconds
        else " Tunggu beberapa saat sebelum mencoba lagi."
    )
    st.warning(
        "🟡 Gemini AI: batas kuota/rate limit tercapai. "
        "Aplikasi tetap berjalan dengan ringkasan fallback."
        f"{wait_text} Periksa penggunaan di "
        "https://ai.dev/rate-limit"
    )


current_year = max(START_YEAR, datetime.now().year)
active_year = int(st.session_state.get("monitor_year", current_year))
active_year = min(max(active_year, START_YEAR), current_year)
st.session_state.monitor_year = active_year
if "data" not in st.session_state or st.session_state.get("data_year") != active_year:
    st.session_state.data = load_data(active_year)
    st.session_state.data_year = active_year
if "gemini_error" not in st.session_state:
    st.session_state.gemini_error = ""

# ============================================================
# STATUS DAN PENGAMBILAN BERITA
# ============================================================
if is_quota_error(st.session_state.gemini_error):
    show_quota_warning(st.session_state.gemini_error)
elif has_valid_api_key_format() and not st.session_state.gemini_error:
    st.success("🟢 Gemini AI: Active")
else:
    if not get_api_key():
        detail = "cek Secrets"
    elif not has_valid_api_key_format():
        detail = "format credential Gemini tidak dikenali"
    else:
        detail = "API gagal; klasifikasi aturan tetap aktif"
    st.error(f"🔴 Gemini AI: Offline - {detail}")

repository = get_repository()
today = datetime.now()

st.markdown('<div class="section-header">Status Pemindaian Berita</div>', unsafe_allow_html=True)
selected_year = st.selectbox(
    "Tahun yang ingin dipantau",
    options=list(range(START_YEAR, current_year + 1)),
    key="monitor_year",
    help=f"Data tersedia mulai {START_YEAR}. Tahun baru otomatis muncul saat kalender berganti.",
)
maximum_month = today.month if selected_year == today.year else 12
selected_month = st.selectbox(
    "Bulan yang ingin dicari",
    options=list(range(1, maximum_month + 1)),
    format_func=lambda month: f"{MONTH_NAMES_ID[month]} {selected_year}",
    help="Pilih satu bulan. Pencarian dan antrean diproses khusus untuk tahun dan bulan tersebut.",
)
candidate_counts = repository.candidate_status_counts(
    year=selected_year,
    month=selected_month,
)
pending_candidates = candidate_counts.get("pending", 0)
rejected_candidates = candidate_counts.get("rejected", 0) + candidate_counts.get("duplicate", 0)
failed_candidates = candidate_counts.get("failed", 0)

news_column, queue_column, rejected_column, month_column = st.columns(4)
news_metric = news_column.empty()
queue_metric = queue_column.empty()
rejected_metric = rejected_column.empty()
month_metric = month_column.empty()
news_metric.metric("Total Berita Lolos", len(st.session_state.data))
queue_metric.metric("Antrean Bulan Dipilih", pending_candidates)
rejected_metric.metric("Ditolak / Duplikat", rejected_candidates)
month_metric.metric("Bulan Dipilih", MONTH_NAMES_ID[selected_month])

if not candidate_counts:
    st.info(
        f"Bulan {MONTH_NAMES_ID[selected_month]} belum pernah dipindai. Menjalankan ulang "
        f"Streamlit hanya memuat {len(st.session_state.data)} data tersimpan; klik tombol "
        "di bawah untuk mulai mencari dan memproses kandidat bulan ini."
    )
elif pending_candidates:
    st.info(
        f"Masih ada {pending_candidates} kandidat dalam antrean. Klik tombol di bawah "
        "untuk memproses batch berikutnya."
    )
elif failed_candidates:
    st.warning(
        f"{failed_candidates} kandidat gagal diambil setelah beberapa percobaan. "
        "Backfill dapat tetap dilanjutkan ke bulan berikutnya."
    )

st.divider()
st.caption(
    "Pemindaian tidak berjalan otomatis saat halaman dimuat ulang. Setiap klik menyimpan "
    "progres ke database dan memproses maksimal 80 kandidat agar tetap stabil."
)

if st.button(
    f"🚀 Cari & Proses Berita {MONTH_NAMES_ID[selected_month]} {selected_year}",
    width="stretch",
    type="primary",
):
    progress_bar = st.progress(0)
    status = st.empty()

    def show_progress(current: int, total: int, message: str) -> None:
        progress_bar.progress(min(current / max(total, 1), 1.0))
        status.text(message)

    with st.spinner("🔎 Mengambil dan menganalisis berita..."):
        classifier = build_classifier()
        pipeline = NewsPipeline(repository, classifier)
        report = pipeline.run(
            show_progress,
            search_year=selected_year,
            search_month=selected_month,
        )
    progress_bar.empty()
    status.empty()
    st.session_state.gemini_error = report.ai_error
    st.session_state.data = load_data(selected_year)
    st.session_state.data_year = selected_year
    updated_counts = repository.candidate_status_counts(
        year=selected_year,
        month=selected_month,
    )
    updated_rejected = updated_counts.get("rejected", 0) + updated_counts.get("duplicate", 0)
    news_metric.metric("Total Berita Lolos", len(st.session_state.data))
    queue_metric.metric("Antrean Bulan Dipilih", updated_counts.get("pending", 0))
    rejected_metric.metric("Ditolak / Duplikat", updated_rejected)
    month_metric.metric("Bulan Dipilih", MONTH_NAMES_ID[selected_month])

    if is_quota_error(report.ai_error):
        show_quota_warning(report.ai_error)

    if report.saved:
        st.success(
            f"✅ Berhasil menyimpan {report.saved} berita ekonomi "
            f"({report.ai_classified} dianalisis Gemini, "
            f"{report.fallback_classified} menggunakan fallback)."
        )
    elif report.processing == 0 and report.discovered == 0:
        st.warning(f"⚠️ Tidak ada berita tahun {selected_year} yang ditemukan dari sumber RSS.")
    elif report.processing == 0:
        st.info(f"ℹ️ Seluruh kandidat {MONTH_NAMES_ID[selected_month]} yang ditemukan sudah diproses.")
    elif report.accepted == 0:
        st.warning("⚠️ Kandidat diproses, tetapi tidak ada berita ekonomi Lamongan yang lolos aturan kualitas.")
    else:
        st.info("ℹ️ Berita yang ditemukan sudah tersimpan sebelumnya.")

    st.caption(
        f"Periode pencarian: {report.search_period or '-'} · "
        f"ditemukan {report.discovered} · diproses {report.processing} · "
        f"gagal ekstraksi {report.extraction_failed} · ditolak {report.rejected} · "
        f"antrean tersisa {report.queued}."
    )

with st.expander("⚙️ Pemeliharaan Data"):
    st.warning(
        f"Reset hanya menghapus berita dan antrean tahun {selected_year}. "
        "Data tahun lain tetap aman."
    )
    confirm_reset = st.checkbox(
        f"Saya memahami bahwa data tahun {selected_year} akan dihapus permanen",
        key=f"confirm_reset_{selected_year}",
    )
    if st.button(
        f"🗑️ Reset Data Tahun {selected_year}",
        width="stretch",
        disabled=not confirm_reset,
    ):
        repository.clear(selected_year)
        st.session_state.data = create_empty_data()
        st.session_state.data_year = selected_year
        st.success(f"✅ Data tahun {selected_year} berhasil direset.")
        st.rerun()

# ============================================================
# FILTER
# ============================================================
df = st.session_state.data.copy()
for column in UI_COLUMNS:
    if column not in df.columns:
        df[column] = ""
df["Tanggal Berita"] = pd.to_datetime(df["Tanggal Berita"], errors="coerce")
df = df[df["Tanggal Berita"].dt.year == selected_year].copy()

with st.sidebar:
    valid_dates = df["Tanggal Berita"].dropna()
    if not valid_dates.empty:
        min_date = valid_dates.min().date()
        max_date = valid_dates.max().date()
    else:
        min_date = datetime(selected_year, 1, 1).date()
        max_date = datetime(selected_year, 12, 31).date()

    date_range = st.date_input("📅 Periode Berita", value=(min_date, max_date))
    selected_media = st.multiselect("🌐 Media", sorted(df["Media"].dropna().unique()))
    selected_sector = st.multiselect("🏭 Sektor Lapangan Usaha", sorted(df["Sektor"].dropna().unique()))
    selected_issue = st.multiselect("📊 Isu Ekonomi", sorted(df["Isu Ekonomi"].dropna().unique()))
    keyword = st.text_input("🔎 Cari kata kunci", placeholder="Ketik kata kunci...")

filtered = df.copy()
if len(date_range) == 2:
    filtered = filtered[
        (filtered["Tanggal Berita"].dt.date >= date_range[0])
        & (filtered["Tanggal Berita"].dt.date <= date_range[1])
    ].copy()
if selected_media:
    filtered = filtered[filtered["Media"].isin(selected_media)].copy()
if selected_sector:
    filtered = filtered[filtered["Sektor"].isin(selected_sector)].copy()
if selected_issue:
    filtered = filtered[filtered["Isu Ekonomi"].isin(selected_issue)].copy()
if keyword:
    search_text = keyword.casefold()
    searchable = filtered[["Judul Berita", "Isu Ekonomi", "Sektor", "Ringkasan Berita"]].fillna("").astype(str)
    mask = searchable.apply(
        lambda row: row.str.casefold().str.contains(search_text, regex=False).any(), axis=1
    )
    filtered = filtered[mask].copy()

# ============================================================
# DASHBOARD UTAMA
# ============================================================
logo_base64 = ""
if LOGO_PATH.exists():
    with open(LOGO_PATH, "rb") as logo_file:
        logo_base64 = base64.b64encode(logo_file.read()).decode("utf-8")
header_img_tag = (
    f'<img src="data:image/png;base64,{logo_base64}" class="dashboard-logo" alt="Logo BPS">'
    if logo_base64
    else ""
)
st.markdown(
    f"""
<div class="dashboard-header">
    {header_img_tag}
    <div>
        <div class="dashboard-title">MONITORING BERITA EKONOMI LAMONGAN</div>
        <div class="dashboard-subtitle">Sistem pemantauan media otomatis berbasis AI untuk 17 Sektor Lapangan Usaha BPS Kabupaten Lamongan</div>
    </div>
</div>
""",
    unsafe_allow_html=True,
)

k1, k2, k3, k4 = st.columns(4)
k1.metric("📰 Total Berita", f"{len(filtered):,}")
k2.metric("📅 Berita Hari Ini", f"{len(filtered[filtered['Tanggal Berita'].dt.date == datetime.now().date()]):,}")
k3.metric("🌐 Sumber Media", f"{filtered['Media'].nunique():,}")
k4.metric("🏭 Sektor Terpantau", f"{filtered['Sektor'].nunique():,}")
st.markdown("<br>", unsafe_allow_html=True)

if not filtered.empty:
    st.markdown('<div class="section-header">📊 Ringkasan Visual & Grafik</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        sector_df = filtered["Sektor"].value_counts().rename_axis("Sektor").reset_index(name="Jumlah")
        fig_sector = px.bar(
            sector_df, x="Jumlah", y="Sektor", orientation="h", text="Jumlah", color="Jumlah",
            color_continuous_scale="Blues", title="Sebaran 17 Sektor Lapangan Usaha BPS",
        )
        fig_sector.update_layout(height=450, showlegend=False, yaxis={"categoryorder": "total ascending"})
        st.plotly_chart(fig_sector, width="stretch")
    with col2:
        media_df = filtered["Media"].value_counts().rename_axis("Media").reset_index(name="Jumlah")
        fig_media = px.pie(media_df, names="Media", values="Jumlah", hole=0.4, title="Proporsi Berita Per Media")
        fig_media.update_layout(height=450)
        st.plotly_chart(fig_media, width="stretch")

    trend_df = filtered.groupby("Tanggal Berita").size().reset_index(name="Jumlah Berita")
    fig_trend = px.area(
        trend_df, x="Tanggal Berita", y="Jumlah Berita", title="📈 Tren Volume Berita Ekonomi",
        color_discrete_sequence=["#2563eb"],
    )
    fig_trend.update_layout(height=300)
    st.plotly_chart(fig_trend, width="stretch")

st.markdown('<div class="section-header">📋 Tabel Berita Terfilter</div>', unsafe_allow_html=True)
if not filtered.empty:
    display_df = filtered.copy()
    display_df["Tanggal Berita"] = display_df["Tanggal Berita"].dt.strftime("%Y-%m-%d")
    st.dataframe(
        display_df[UI_COLUMNS],
        column_config={
            "Link Berita": st.column_config.LinkColumn("Link Berita", display_text="🔗 Baca Artikel"),
            "Ringkasan Berita": st.column_config.TextColumn("Ringkasan Berita", width="large"),
        },
        width="stretch",
        hide_index=True,
    )
else:
    st.warning("Tidak ada data berita yang cocok dengan filter.")

st.markdown('<div class="section-header">📥 Ekspor Laporan Excel</div>', unsafe_allow_html=True)
if not filtered.empty:
    export_df = filtered.copy()
    export_df["Tanggal Berita"] = export_df["Tanggal Berita"].dt.strftime("%Y-%m-%d")
    export_df = export_df[UI_COLUMNS]
    try:
        from openpyxl.styles import Alignment, Font, PatternFill

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Monitoring Berita")
            worksheet = writer.sheets["Monitoring Berita"]
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            for col_num in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_letter, width in {"A": 18, "B": 20, "C": 35, "D": 22, "E": 38, "F": 50, "G": 30}.items():
                worksheet.column_dimensions[col_letter].width = width
            body_alignment = Alignment(vertical="top", wrap_text=True)
            for row in worksheet.iter_rows(
                min_row=2, max_row=len(export_df) + 1, min_col=1, max_col=len(export_df.columns)
            ):
                for cell in row:
                    cell.alignment = body_alignment
        buffer.seek(0)
        st.download_button(
            label="📊 Download Laporan Excel (.xlsx)",
            data=buffer,
            file_name=f"Laporan_Berita_Ekonomi_Lamongan_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    except ImportError:
        st.info("💡 Pastikan 'openpyxl' sudah ada di requirements.txt")

# ============================================================
# TEST GEMINI AI (developer utility)
# ============================================================
with st.expander("🧪 Test Gemini AI"):
    st.caption("Gunakan panel ini untuk memverifikasi koneksi dan respons model Gemini secara manual.")
    test_title = st.text_input(
        "Judul Berita",
        value="Harga cabai di Lamongan mengalami kenaikan",
        key="test_title",
    )
    test_content = st.text_area(
        "Isi Berita",
        value=(
            "Harga cabai rawit di sejumlah pasar Kabupaten Lamongan "
            "mengalami kenaikan akibat berkurangnya pasokan dari petani. "
            "Kenaikan harga tersebut berdampak terhadap pedagang dan "
            "konsumen di wilayah Lamongan."
        ),
        height=180,
        key="test_content",
    )
    if st.button("🤖 Test Analisis Gemini", width="stretch", key="btn_test_gemini"):
        if not test_content.strip():
            st.warning("Isi berita harus diisi.")
        else:
            classifier = build_classifier()
            with st.spinner("Gemini sedang menganalisis..."):
                result = classifier.test(test_title, test_content)
            st.session_state.gemini_error = classifier.last_error
            st.write("### Hasil Analisis")
            st.json(result)

st.divider()
st.caption("Dashboard Monitoring Berita Ekonomi Kabupaten Lamongan | BPS Kabupaten Lamongan")
